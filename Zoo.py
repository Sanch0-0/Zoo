from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class Zoo:
    def __init__(self):
        self.__animal_list = []
        # (Образ питания)
        self._carnivorous_list = []
        self._herbivorous_list = []
        self._omnivorous_list = []
        # (Тип животного)
        self._mammal_list = []
        self._reptile_list = []
        self._artiodactyl_list = []
        self._bird_list = []
    
    def add_animals(self, *animals):
        for animal in animals:
            self.__animal_list.append(animal)

            if animal._animal_type == "Млекопитающее":
                self._mammal_list.append(animal)
            elif animal._animal_type == "Рептилия":
                self._reptile_list.append(animal)
            elif animal._animal_type == "Парнокопытное":
                self._artiodactyl_list.append(animal)
            else:
                self._bird_list.append(animal)

            if animal._food_type == "Травоядное":
                self._herbivorous_list.append(animal)
            elif animal._food_type == "Плотоядное":
                self._carnivorous_list.append(animal)
            else:
                self._omnivorous_list.append(animal)

    def remove_animals(self, *animals):
        for animal in animals:
            if animal in self.__animal_list:
                if animal in self._mammal_list:
                    self._mammal_list.remove(animal)
                elif animal in self._reptile_list:
                    self._reptile_list.remove(animal)
                elif animal in self._artiodactyl_list:
                    self._artiodactyl_list.remove(animal)
                else:
                    self._bird_list.remove(animal)

                if animal in self._carnivorous_list:
                    self._carnivorous_list.remove(animal)
                elif animal in self._herbivorous_list:
                    self._herbivorous_list.remove(animal)
                else:
                    self._omnivorous_list.remove(animal)

                self.__animal_list.remove(animal)

    def feed_all(self):
        for animal in self.__animal_list:
            ration = animal._calculate_daily_ration()
            print(
                f"{animal.name} получил {ration:.2f}кг корма."
            )

    def show_total_info(self):
        for animal in self.__animal_list:
            animal.show_info()

    def make_all_sounds(self):
        for animal in self.__animal_list:
            animal.make_sound()

    # def __count_animal_category(self):
    #     print(f'''
    #     Всего животных: {len(self.__animal_list)};
    #     Всего Млекопитающих: {len(self._mammal_list)};
    #     Всего Рептилий: {len(self._reptile_list)};
    #     Всего Парнокопытных: {len(self._artiodactyl_list)};
    #     Всего Птиц: {len(self._bird_list)}.
    #     ''')

    # def __count_food_category(self):
    #     print(f'''
    #     Всего Плотоядных животных: {len(self._carnivorous_list)};
    #     Всего Травоядных животных: {len(self._herbivorous_list)}; 
    #     Всего Всеядных животных: {len(self._omnivorous_list)}.
    #     ''')

    # def __show_weight_stats(self):
    #     haviest_mammal = max(self._mammal_list, key=lambda animal: animal._weight)
    #     haviest_reptile = max(self._reptile_list, key=lambda animal: animal._weight)
    #     haviest_artiodactyl = max(self._artiodactyl_list, key=lambda animal: animal._weight)
    #     haviest_bird = max(self._bird_list, key=lambda animal: animal._weight)

    #     lightest_mammal = min(self._mammal_list, key=lambda animal: animal._weight)
    #     lightest_reptile = min(self._reptile_list, key=lambda animal: animal._weight)
    #     lightest_artiodactyl = min(self._artiodactyl_list, key=lambda animal: animal._weight)
    #     lightest_bird = min(self._bird_list, key=lambda animal: animal._weight)

    #     the_lightest = min(self.__animal_list, key=lambda animal: animal._weight)
    #     the_heaviest = max(self.__animal_list, key=lambda animal: animal._weight)

    #     print(f'''
    #     Общий вес Млекопитающих: {sum(animal._weight for animal in self._mammal_list)} кг;
    #     Общий вес Рептилий: {sum(animal._weight for animal in self._reptile_list)} кг;
    #     Общий вес Парнокопытных: {sum(animal._weight for animal in self._artiodactyl_list)} кг;
    #     Общий вес Птиц: {sum(animal._weight for animal in self._bird_list)} кг;

    #     Средний вес Млекопитающих: {sum(animal._weight for animal in self._mammal_list) / len(self._mammal_list):.2f} кг;
    #     Средний вес Рептилий: {sum(animal._weight for animal in self._reptile_list) / len(self._reptile_list):.2f} кг;
    #     Средний вес Парнокопытных: {sum(animal._weight for animal in self._artiodactyl_list) / len(self._artiodactyl_list):.2f} кг;
    #     Средний вес Птиц: {sum(animal._weight for animal in self._bird_list) / len(self._bird_list):.2f} кг;

    #     Самое тяжёлое Млекопитающее: {haviest_mammal.name} - {haviest_mammal._weight} кг;
    #     Самая тяжёлая Рептилия: {haviest_reptile.name} - {haviest_reptile._weight} кг;
    #     Самое тяжёлое Парнокопытное: {haviest_artiodactyl.name} - {haviest_artiodactyl._weight} кг;
    #     Самая тяжёлая Птица: {haviest_bird.name} - {haviest_bird._weight} кг;

    #     Самое легкое Млекопитающее: {lightest_mammal.name} - {lightest_mammal._weight} кг;
    #     Самая легкая Рептилия: {lightest_reptile.name} - {lightest_reptile._weight} кг;
    #     Самое легкое Парнокопытное: {lightest_artiodactyl.name} - {lightest_artiodactyl._weight} кг;
    #     Самая легкая Птица: {lightest_bird.name} - {lightest_bird._weight} кг;

    #     Легчайшее животное: {the_lightest.name} - {the_lightest._weight} кг;
    #     Тяжелейшее животное: {the_heaviest.name} - {the_heaviest._weight} кг.
    #     ''')

    # def __show_age_stats(self):
    #     oldest_mammal = max(self._mammal_list, key=lambda animal: animal.age)
    #     oldest_reptile = max(self._reptile_list, key=lambda animal: animal.age)
    #     oldest_artiodactyl = max(self._artiodactyl_list, key=lambda animal: animal.age)
    #     oldest_bird = max(self._bird_list, key=lambda animal: animal.age)

    #     youngest_mammal = min(self._mammal_list, key=lambda animal: animal.age)
    #     youngest_reptile = min(self._reptile_list, key=lambda animal: animal.age)
    #     youngest_artiodactyl = min(self._artiodactyl_list, key=lambda animal: animal.age)
    #     youngest_bird = min(self._bird_list, key=lambda animal: animal.age)

    #     the_youngest = min(self.__animal_list, key=lambda animal: animal.age)
    #     the_oldest = max(self.__animal_list, key=lambda animal: animal.age)

    #     print(f'''
    #     Средний возраст Млекопитающих: {sum(animal.age for animal in self._mammal_list) / len(self._mammal_list):.2f} лет;
    #     Средний возраст Рептилий: {sum(animal.age for animal in self._reptile_list) / len(self._reptile_list):.2f} лет;
    #     Средний возраст Парнокопытных: {sum(animal.age for animal in self._artiodactyl_list) / len(self._artiodactyl_list):.2f} лет;
    #     Средний возраст Птиц: {sum(animal.age for animal in self._bird_list) / len(self._bird_list):.2f} лет;

    #     Самое старое Млекопитающее: {oldest_mammal.name} - {oldest_mammal.age} лет;
    #     Самая старая Рептилия: {oldest_reptile.name} - {oldest_reptile.age} лет;
    #     Самое старое Парнокопытное: {oldest_artiodactyl.name} - {oldest_artiodactyl.age}лет;
    #     Самая старая Птица: {oldest_bird.name} - {oldest_bird.age} лет;

    #     Самое молодое Млекопитающее: {youngest_mammal.name} - {youngest_mammal.age} лет;
    #     Самая молодая Рептилия: {youngest_reptile.name} - {youngest_reptile.age} лет;
    #     Самое молодое Парнокопытное: {youngest_artiodactyl.name} - {youngest_artiodactyl.age}лет;
    #     Самая молодая Птица: {youngest_bird.name} - {youngest_bird.age} лет;

    #     Cтарейщее животное: {the_oldest.name} - {the_oldest.age} лет;
    #     Самое молодое животное: {the_youngest.name} - {the_youngest.age} лет.
    #     ''')

    # def __show_food_stats(self):
    #     the_most_greedy = max(self.__animal_list, key=lambda animal: animal._daily_ration)
    #     the_least_greedy = min(self.__animal_list, key=lambda animal: animal._daily_ration)

    #     print(f'''
    #     Всего корма для плотоядных: {sum(animal._daily_ration for animal in self._carnivorous_list):.2f} кг;
    #     Всего корма для травоядных: {sum(animal._daily_ration for animal in self._herbivorous_list)} кг;
    #     Всего корма для всеядных: {sum(animal._daily_ration for animal in self._omnivorous_list)} кг;

    #     Больше всего корма нужно для: {the_most_greedy.name} - {the_most_greedy._daily_ration} кг;
    #     Меньше всего корма нужно для: {the_least_greedy.name} - {the_least_greedy._daily_ration} кг;

    #     Общее кол-во корма для всех животных: {sum(animal._daily_ration for animal in self.__animal_list)} кг.
    #     ''')


    # def generate_report(self):
    #     # TODO - rebuild that shit into pandas -> excel file output!
    #     self.__count_animal_category()
    #     self.__count_food_category()
    #     self.__show_weight_stats()
    #     self.__show_age_stats()
    #     self.__show_food_stats()


    def __count_animal_category(self):
        return [
            ["Всего животных", len(self.__animal_list)],
            ["Всего Млекопитающих", len(self._mammal_list)],
            ["Всего Рептилий", len(self._reptile_list)],
            ["Всего Парнокопытных", len(self._artiodactyl_list)],
            ["Всего Птиц", len(self._bird_list)],
        ]


    def __count_food_category(self):
        return [
            ["Всего Плотоядных животных", len(self._carnivorous_list)],
            ["Всего Травоядных животных", len(self._herbivorous_list)],
            ["Всего Всеядных животных", len(self._omnivorous_list)],
        ]


    def __show_weight_stats(self):
        haviest_mammal = max(self._mammal_list, key=lambda animal: animal._weight)
        haviest_reptile = max(self._reptile_list, key=lambda animal: animal._weight)
        haviest_artiodactyl = max(self._artiodactyl_list, key=lambda animal: animal._weight)
        haviest_bird = max(self._bird_list, key=lambda animal: animal._weight)

        lightest_mammal = min(self._mammal_list, key=lambda animal: animal._weight)
        lightest_reptile = min(self._reptile_list, key=lambda animal: animal._weight)
        lightest_artiodactyl = min(self._artiodactyl_list, key=lambda animal: animal._weight)
        lightest_bird = min(self._bird_list, key=lambda animal: animal._weight)

        the_lightest = min(self.__animal_list, key=lambda animal: animal._weight)
        the_heaviest = max(self.__animal_list, key=lambda animal: animal._weight)

        return [
        ["Общий вес Млекопитающих:", f"{sum(animal._weight for animal in self._mammal_list)} кг"],
        ["Общий вес Рептилий:", f"{sum(animal._weight for animal in self._reptile_list)} кг"],
        ["Общий вес Парнокопытных:", f"{sum(animal._weight for animal in self._artiodactyl_list)} кг"],
        ["Общий вес Птиц:", f"{sum(animal._weight for animal in self._bird_list)} кг"],

        ["Средний вес Млекопитающих:", f"{sum(animal._weight for animal in self._mammal_list) / len(self._mammal_list):.2f} кг"],
        ["Средний вес Рептилий:", f"{sum(animal._weight for animal in self._reptile_list) / len(self._reptile_list):.2f} кг"],
        ["Средний вес Парнокопытных:", f"{sum(animal._weight for animal in self._artiodactyl_list) / len(self._artiodactyl_list):.2f} кг"],
        ["Средний вес Птиц:", f"{sum(animal._weight for animal in self._bird_list) / len(self._bird_list):.2f} кг"],

        ["Самое тяжёлое Млекопитающее:", f"{haviest_mammal.name} - {haviest_mammal._weight} кг"],
        ["Самая тяжёлая Рептилия:", f"{haviest_reptile.name} - {haviest_reptile._weight} кг"],
        ["Самое тяжёлое Парнокопытное:", f"{haviest_artiodactyl.name} - {haviest_artiodactyl._weight} кг"],
        ["Самая тяжёлая Птица:", f"{haviest_bird.name} - {haviest_bird._weight} кг"],

        ["Самое легкое Млекопитающее:", f"{lightest_mammal.name} - {lightest_mammal._weight} кг"],
        ["Самая легкая Рептилия:", f"{lightest_reptile.name} - {lightest_reptile._weight} кг"],
        ["Самое легкое Парнокопытное:", f"{lightest_artiodactyl.name} - {lightest_artiodactyl._weight} кг"],
        ["Самая легкая Птица:", f"{lightest_bird.name} - {lightest_bird._weight} кг"],

        ["Легчайшее животное:", f"{the_lightest.name} - {the_lightest._weight} кг"],
        ["Тяжелейшее животное:", f"{the_heaviest.name} - {the_heaviest._weight} кг"]
        ]


    def __show_age_stats(self):
        oldest_mammal = max(self._mammal_list, key=lambda animal: animal.age)
        oldest_reptile = max(self._reptile_list, key=lambda animal: animal.age)
        oldest_artiodactyl = max(self._artiodactyl_list, key=lambda animal: animal.age)
        oldest_bird = max(self._bird_list, key=lambda animal: animal.age)

        youngest_mammal = min(self._mammal_list, key=lambda animal: animal.age)
        youngest_reptile = min(self._reptile_list, key=lambda animal: animal.age)
        youngest_artiodactyl = min(self._artiodactyl_list, key=lambda animal: animal.age)
        youngest_bird = min(self._bird_list, key=lambda animal: animal.age)

        the_youngest = min(self.__animal_list, key=lambda animal: animal.age)
        the_oldest = max(self.__animal_list, key=lambda animal: animal.age)

        return [
        ["Средний возраст Млекопитающих:", f"{sum(animal.age for animal in self._mammal_list) / len(self._mammal_list):.1f} лет"],
        ["Средний возраст Рептилий:", f"{sum(animal.age for animal in self._reptile_list) / len(self._reptile_list):.1f} лет"],
        ["Средний возраст Парнокопытных:", f"{sum(animal.age for animal in self._artiodactyl_list) / len(self._artiodactyl_list):.1f} лет"],
        ["Средний возраст Птиц:", f"{sum(animal.age for animal in self._bird_list) / len(self._bird_list):.1f} лет"],

        ["Самое старое Млекопитающее:", f"{oldest_mammal.name} - {oldest_mammal.age} лет"],
        ["Самая старая Рептилия:", f"{oldest_reptile.name} - {oldest_reptile.age} лет"],
        ["Самое старое Парнокопытное:", f"{oldest_artiodactyl.name} - {oldest_artiodactyl.age} лет"],
        ["Самая старая Птица:", f"{oldest_bird.name} - {oldest_bird.age} лет"],

        ["Самое молодое Млекопитающее:", f"{youngest_mammal.name} - {youngest_mammal.age} лет"],
        ["Самая молодая Рептилия:", f"{youngest_reptile.name} - {youngest_reptile.age} лет"],
        ["Самое молодое Парнокопытное:", f"{youngest_artiodactyl.name} - {youngest_artiodactyl.age}лет"],
        ["Самая молодая Птица:", f"{youngest_bird.name} - {youngest_bird.age} лет"],

        ["Cтарейщее животное:", f"{the_oldest.name} - {the_oldest.age} лет"],
        ["Самое молодое животное:", f"{the_youngest.name} - {the_youngest.age} лет"]
        ]


    def __show_food_stats(self):
        the_most_greedy = max(self.__animal_list, key=lambda animal: animal._daily_ration)
        the_least_greedy = min(self.__animal_list, key=lambda animal: animal._daily_ration)

        return [
            ["Всего корма для плотоядных", f"{sum(animal._daily_ration for animal in self._carnivorous_list):.2f} кг"],
            ["Всего корма для травоядных", f"{sum(animal._daily_ration for animal in self._herbivorous_list):.2f} кг"],
            ["Всего корма для всеядных", f"{sum(animal._daily_ration for animal in self._omnivorous_list):.2f} кг"],
            ["Больше всего корма нужно для", f"{the_most_greedy.name} - {the_most_greedy._daily_ration} кг"],
            ["Меньше всего корма нужно для", f"{the_least_greedy.name} - {the_least_greedy._daily_ration} кг"],
        ]

    def generate_report(self):
        # Создание нового Excel файла
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Параметр"
        ws['B1'] = "Значение"
        ws.title = "Zoo Report"

        # Данные для отчета (каждый метод возвращает список строк)
        report_sections = [
            ("Общее количество", self.__count_animal_category()),
            ("Образ питания", self.__count_food_category()),
            ("Вес", self.__show_weight_stats()),
            ("Возраст", self.__show_age_stats()),
            ("Корм", self.__show_food_stats()),
        ]

        # Настройки для форматирования
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True)
        alignment = Alignment(horizontal="left", vertical="top")
        row_number = 2

        # Добавление данных в один лист
        for section_name, data in report_sections:
            # Заголовок секции
            ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
            header_cell = ws.cell(row=row_number, column=1, value=section_name)
            header_cell.font = header_font
            header_cell.alignment = alignment
            row_number += 1

            # Добавление данных секции
            for row in data:
                for col_index, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_number, column=col_index, value=value)
                    if row_number == 2:  # Первая строка данных (заголовки)
                        cell.font = subheader_font
                    cell.alignment = alignment
                row_number += 1

            # Пустая строка между секциями
            row_number += 1

        # Сохранение файла
        file_name = "zoo_report.xlsx"
        wb.save(file_name)
        print(f"Отчет успешно сохранен как {file_name}")
